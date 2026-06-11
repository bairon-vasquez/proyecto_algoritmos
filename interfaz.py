"""
Interfaz gráfica GeoMIP — KQNodes / KGeoMIP
Análisis de k-Particiones Óptimas — ADA 2026-1
"""
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import time
import os
import csv
import sys
import queue
import concurrent.futures
from math import comb, factorial
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

try:
    from src.models.system import System
    from src.controllers.strategies.geometric import KGeoMIP, KGeoMIPKPartition
    from src.controllers.strategies.qnodes import KQNodes
    MODULES_OK = True
    MODULES_ERROR = None
except Exception as _exc:
    MODULES_OK = False
    MODULES_ERROR = str(_exc)

try:
    import customtkinter as ctk
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")
    _CTK = True
except ImportError:
    _CTK = False

try:
    import sv_ttk
    _SV_TTK = True
except ImportError:
    _SV_TTK = False

# ── Paleta ────────────────────────────────────────────────────────────────────
PALETTE = {
    "bg":      "#1e1e2e",
    "panel":   "#2a2a3d",
    "card":    "#313244",
    "accent":  "#cba6f7",
    "green":   "#a6e3a1",
    "red":     "#f38ba8",
    "yellow":  "#f9e2af",
    "blue":    "#89b4fa",
    "text":    "#cdd6f4",
    "subtext": "#9399b2",
    "border":  "#45475a",
}

# ── Predefinidos ──────────────────────────────────────────────────────────────
PREDEFINIDOS = {
    "N=10": {"sistema": "ABCDEFGHIJ",              "estado": "1000000000",            "csv": "data/N10C.csv"},
    "N=15": {"sistema": "ABCDEFGHIJKLMNO",          "estado": "100000000000000",       "csv": "data/N15C.csv"},
    "N=20": {"sistema": "ABCDEFGHIJKLMNOPQRST",     "estado": "10000000000000000000",  "csv": "data/N20C.csv"},
    "N=22": {"sistema": "ABCDEFGHIJKLMNOPQRSTUV",   "estado": "1000000000000000000000","csv": "data/N22C.csv"},
    "N=25": {"sistema": "ABCDEFGHIJKLMNOPQRSTUVWXY","estado": "1000000000000000000000000","csv": "data/N25C.csv"},
}

HOJAS_EXCEL = {10: '10A-Elementos', 15: '15B-Elementos', 20: '20A-Elementos',
               22: '22A-Elementos', 25: '25A-Elementos'}

SUITES = {
    'N=10': {
        'sistema': 'ABCDEFGHIJ', 'estado': '1000000000', 'csv': 'data/N10C.csv',
        'pruebas': [
            ('ABCDEFGHIJ','ABCDEFGHIJ'),('ABCDEFGHIJ','ABCDEFGHI'),
            ('ABCDEFGHIJ','BCDEFGHIJ'), ('ABCDEFGHIJ','BCDEFGHI'),
            ('ABCDEFGHIJ','ABDEGHJ'),   ('ABCDEFGHIJ','ACEGI'),
            ('ABCDEFGHIJ','BDFHJ'),     ('ABCDEFGHI', 'ABCDEFGHIJ'),
            ('ABCDEFGHI', 'ABCDEFGHI'), ('ABCDEFGHI', 'BCDEFGHIJ'),
            ('ABCDEFGHI', 'BCDEFGHI'),  ('ABCDEFGHI', 'ABDEGHJ'),
            ('ABCDEFGHI', 'ACEGI'),     ('ABCDEFGHI', 'BDFHJ'),
            ('BCDEFGHIJ', 'ABCDEFGHIJ'),('BCDEFGHIJ', 'ABCDEFGHI'),
            ('BCDEFGHIJ', 'BCDEFGHIJ'), ('BCDEFGHIJ', 'BCDEFGHI'),
            ('BCDEFGHIJ', 'ABDEGHJ'),   ('BCDEFGHIJ', 'ACEGI'),
            ('BCDEFGHIJ', 'BDFHJ'),     ('BCDEFGHI',  'ABCDEFGHIJ'),
            ('BCDEFGHI',  'ABCDEFGHI'), ('BCDEFGHI',  'BCDEFGHIJ'),
            ('BCDEFGHI',  'BCDEFGHI'),  ('BCDEFGHI',  'ABDEGHJ'),
            ('BCDEFGHI',  'ACEGI'),     ('BCDEFGHI',  'BDFHJ'),
            ('ABDEGHJ',   'ABCDEFGHIJ'),('ABDEGHJ',   'ABCDEFGHI'),
            ('ABDEGHJ',   'BCDEFGHIJ'), ('ABDEGHJ',   'BCDEFGHI'),
            ('ABDEGHJ',   'ABDEGHJ'),   ('ABDEGHJ',   'ACEGI'),
            ('ABDEGHJ',   'BDFHJ'),     ('ACEGI',     'ABCDEFGHIJ'),
            ('ACEGI',     'ABCDEFGHI'), ('ACEGI',     'BCDEFGHIJ'),
            ('ACEGI',     'BCDEFGHI'),  ('ACEGI',     'ABDEGHJ'),
            ('ACEGI',     'ACEGI'),     ('ACEGI',     'BDFHJ'),
            ('BDFHJ',     'ABCDEFGHIJ'),('BDFHJ',     'ABCDEFGHI'),
            ('BDFHJ',     'BCDEFGHIJ'), ('BDFHJ',     'BCDEFGHI'),
            ('BDFHJ',     'ABDEGHJ'),   ('BDFHJ',     'ACEGI'),
            ('BDFHJ',     'BDFHJ'),
        ],
    },
    'N=15': {
        'sistema': 'ABCDEFGHIJKLMNO', 'estado': '100000000000000', 'csv': 'data/N15C.csv',
        'pruebas': [
            ('ABCDEFGHIJKLMNO','ABCDEFGHIJKLMNO'),('ABCDEFGHIJKLMNO','ABCDEFGHIJKLMN'),
            ('ABCDEFGHIJKLMNO','BCDEFGHIJKLMNO'), ('ABCDEFGHIJKLMNO','BCDEFGHIJKLMN'),
            ('ABCDEFGHIJKLMNO','ABDEGHJKMN'),      ('ABCDEFGHIJKLMNO','ACEGIKMO'),
            ('ABCDEFGHIJKLMNO','BDFHJLN'),         ('ABCDEFGHIJKLMN', 'ABCDEFGHIJKLMNO'),
            ('ABCDEFGHIJKLMN', 'ABCDEFGHIJKLMN'), ('ABCDEFGHIJKLMN', 'BCDEFGHIJKLMNO'),
            ('ABCDEFGHIJKLMN', 'BCDEFGHIJKLMN'),  ('ABCDEFGHIJKLMN', 'ABDEGHJKMN'),
            ('ABCDEFGHIJKLMN', 'ACEGIKMO'),        ('ABCDEFGHIJKLMN', 'BDFHJLN'),
            ('BCDEFGHIJKLMNO', 'ABCDEFGHIJKLMNO'),('BCDEFGHIJKLMNO', 'ABCDEFGHIJKLMN'),
            ('BCDEFGHIJKLMNO', 'BCDEFGHIJKLMNO'), ('BCDEFGHIJKLMNO', 'BCDEFGHIJKLMN'),
            ('BCDEFGHIJKLMNO', 'ABDEGHJKMN'),      ('BCDEFGHIJKLMNO', 'ACEGIKMO'),
            ('BCDEFGHIJKLMNO', 'BDFHJLN'),         ('BCDEFGHIJKLMN',  'ABCDEFGHIJKLMNO'),
            ('BCDEFGHIJKLMN',  'ABCDEFGHIJKLMN'), ('BCDEFGHIJKLMN',  'BCDEFGHIJKLMNO'),
            ('BCDEFGHIJKLMN',  'BCDEFGHIJKLMN'),  ('BCDEFGHIJKLMN',  'ABDEGHJKMN'),
            ('BCDEFGHIJKLMN',  'ACEGIKMO'),        ('BCDEFGHIJKLMN',  'BDFHJLN'),
            ('ABDEGHJKMN',     'ABCDEFGHIJKLMNO'),('ABDEGHJKMN',     'ABCDEFGHIJKLMN'),
            ('ABDEGHJKMN',     'BCDEFGHIJKLMNO'), ('ABDEGHJKMN',     'BCDEFGHIJKLMN'),
            ('ABDEGHJKMN',     'ABDEGHJKMN'),      ('ABDEGHJKMN',     'ACEGIKMO'),
            ('ABDEGHJKMN',     'BDFHJLN'),         ('ACEGIKMO',       'ABCDEFGHIJKLMNO'),
            ('ACEGIKMO',       'ABCDEFGHIJKLMN'), ('ACEGIKMO',       'BCDEFGHIJKLMNO'),
            ('ACEGIKMO',       'BCDEFGHIJKLMN'),  ('ACEGIKMO',       'ABDEGHJKMN'),
            ('ACEGIKMO',       'ACEGIKMO'),        ('ACEGIKMO',       'BDFHJLN'),
            ('BDFHJLN',        'ABCDEFGHIJKLMNO'),('BDFHJLN',        'ABCDEFGHIJKLMN'),
            ('BDFHJLN',        'BCDEFGHIJKLMNO'), ('BDFHJLN',        'BCDEFGHIJKLMN'),
            ('BDFHJLN',        'ABDEGHJKMN'),      ('BDFHJLN',        'ACEGIKMO'),
            ('BDFHJLN',        'BDFHJLN'),         ('BCDEFGJKLMNO',   'BCDEFGHIJKLMNO'),
        ],
    },
}

def stirling2(n: int, k: int) -> int:
    if k == 0: return 1 if n == 0 else 0
    if k > n:  return 0
    return sum((-1)**(k-j)*comb(k,j)*(j**n) for j in range(k+1)) // factorial(k)


# ── Widgets auxiliares ────────────────────────────────────────────────────────

class MetricCard(tk.Frame):
    def __init__(self, parent, title, **kw):
        P = PALETTE
        super().__init__(parent, bg=P["card"], relief=tk.FLAT,
                         highlightthickness=1, highlightbackground=P["border"], **kw)
        tk.Label(self, text=title, bg=P["card"], fg=P["subtext"],
                 font=("Segoe UI", 8)).pack(anchor="w", padx=10, pady=(8,2))
        self._val = tk.StringVar(value="—")
        self._lbl = tk.Label(self, textvariable=self._val, bg=P["card"],
                             fg=P["accent"], font=("Segoe UI", 13, "bold"))
        self._lbl.pack(anchor="w", padx=10, pady=(0,8))

    def set(self, value, color=None):
        self._val.set(str(value))
        if color:
            self._lbl.configure(fg=color)


class ColoredLog(tk.Text):
    _TAGS = {
        "ok":   ("#a6e3a1", None),
        "err":  ("#f38ba8", None),
        "warn": ("#f9e2af", None),
        "info": ("#89b4fa", None),
        "head": ("#cba6f7", "bold"),
    }
    def __init__(self, parent, **kw):
        P = PALETTE
        defaults = dict(bg=P["bg"], fg=P["text"], font=("Consolas", 9),
                        relief=tk.FLAT, state=tk.DISABLED, wrap=tk.WORD,
                        selectbackground=P["border"])
        defaults.update(kw)
        super().__init__(parent, **defaults)
        for tag, (fg, weight) in self._TAGS.items():
            font_spec = ("Consolas", 9, weight) if weight else ("Consolas", 9)
            self.tag_configure(tag, foreground=fg, font=font_spec)

    def append(self, text: str, tag: str = ""):
        self.configure(state=tk.NORMAL)
        self.insert(tk.END, text + "\n", (tag,) if tag else ())
        self.see(tk.END)
        self.configure(state=tk.DISABLED)

    def clear(self):
        self.configure(state=tk.NORMAL)
        self.delete("1.0", tk.END)
        self.configure(state=tk.DISABLED)

    def set_content(self, text: str):
        self.configure(state=tk.NORMAL)
        self.delete("1.0", tk.END)
        self.insert("1.0", text)
        self.configure(state=tk.DISABLED)


class ResultsTable(ttk.Treeview):
    _COLS = ("Estrategia", "k", "Partición", "φ", "Tiempo (s)")

    def __init__(self, parent, **kw):
        super().__init__(parent, columns=self._COLS, show="headings",
                         selectmode="browse", **kw)
        widths = [110, 40, 260, 90, 80]
        for col, w in zip(self._COLS, widths):
            self.heading(col, text=col)
            self.column(col, width=w, minwidth=40, anchor="w")
        P = PALETTE
        self.tag_configure("phi_zero",    foreground=P["green"])
        self.tag_configure("phi_nonzero", foreground=P["yellow"])
        self.tag_configure("timeout_row", foreground=P["red"])

    def add_row(self, values):
        try:
            phi_val = float(values[3])
            tag = "phi_zero" if phi_val < 1e-9 else "phi_nonzero"
        except (ValueError, TypeError, IndexError):
            tag = "timeout_row"
        self.insert("", tk.END, values=values, tags=(tag,))

    def clear(self):
        for item in self.get_children():
            self.delete(item)


# ── App principal ─────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        P = PALETTE
        self.title("GeoMIP — k-Particiones Óptimas  (ADA 2026-1)")
        self.geometry("1440x860")
        self.minsize(1100, 700)
        self.configure(bg=P["bg"])

        if _SV_TTK:
            sv_ttk.set_theme("dark")
        else:
            style = ttk.Style(self)
            try:
                style.theme_use("clam")
            except tk.TclError:
                pass
            style.configure("Treeview", background=P["card"], foreground=P["text"],
                             fieldbackground=P["card"], rowheight=22, font=("Consolas", 9))
            style.configure("Treeview.Heading", background=P["panel"],
                             foreground=P["subtext"], font=("Segoe UI", 9, "bold"))
            style.map("Treeview", background=[("selected", P["border"])])
            style.configure("TProgressbar", troughcolor=P["panel"],
                             background=P["accent"], thickness=6)

        # Estado interno
        self._q               = queue.Queue()
        self._results         = []
        self._historial_analisis = []
        self._ultimo_res_geo  = None
        self._ultimo_res_qn   = None
        self._ultimo_sub      = None
        self._ultimo_t_geo    = 0.0
        self._ultimo_t_qn     = 0.0
        self._ultimo_n_sistema= 0
        self._ultimo_sistema  = ""
        self._modo_suite_activo = False
        self._suite_running   = False

        self._build_ui()
        self.after(80, self._poll)

        if not MODULES_OK:
            self.after(200, lambda: messagebox.showwarning(
                "Módulos no cargados",
                f"Error importando módulos del proyecto:\n{MODULES_ERROR}\n\n"
                "Verifica que el PYTHONPATH y las dependencias estén instaladas."))

    # ── Construcción de UI ────────────────────────────────────────────────────

    def _build_ui(self):
        P = PALETTE
        # Separador superior decorativo
        tk.Frame(self, height=3, bg=P["accent"]).pack(fill=tk.X)

        content = tk.Frame(self, bg=P["bg"])
        content.pack(fill=tk.BOTH, expand=True, padx=0, pady=0)

        self._build_sidebar(content)
        self._build_main(content)
        self._build_statusbar()

    def _build_sidebar(self, parent):
        P = PALETTE
        side = tk.Frame(parent, bg=P["panel"], width=300)
        side.pack(side=tk.LEFT, fill=tk.Y, padx=(8, 4), pady=8)
        side.pack_propagate(False)

        # Logo
        hdr = tk.Frame(side, bg=P["panel"])
        hdr.pack(fill=tk.X, padx=12, pady=(14, 4))
        tk.Label(hdr, text="GeoMIP", bg=P["panel"], fg=P["accent"],
                 font=("Segoe UI", 18, "bold")).pack(anchor="w")
        tk.Label(hdr, text="k-Particiones Óptimas  ·  ADA 2026-1",
                 bg=P["panel"], fg=P["subtext"], font=("Segoe UI", 8)).pack(anchor="w")
        tk.Frame(side, height=1, bg=P["border"]).pack(fill=tk.X, padx=8, pady=6)

        def row(label, widget_fn):
            f = tk.Frame(side, bg=P["panel"])
            f.pack(fill=tk.X, padx=12, pady=2)
            tk.Label(f, text=label, bg=P["panel"], fg=P["subtext"],
                     font=("Segoe UI", 8), width=12, anchor="w").pack(side=tk.LEFT)
            widget_fn(f)

        # Predefinido selector
        self._pre_var = tk.StringVar(value="N=10")
        row("Sistema", lambda f: ttk.Combobox(
            f, textvariable=self._pre_var,
            values=list(PREDEFINIDOS.keys()), state="readonly", width=16
        ).pack(side=tk.LEFT))
        self._pre_var.trace_add("write", lambda *_: self._on_predefinido())

        # CSV
        self._csv_var = tk.StringVar(value="data/N10C.csv")
        def _csv_row(f):
            e = tk.Entry(f, textvariable=self._csv_var, width=14,
                         bg=P["card"], fg=P["text"], insertbackground=P["text"],
                         relief=tk.FLAT, font=("Consolas", 9))
            e.pack(side=tk.LEFT, padx=(0, 4))
            tk.Button(f, text="…", command=self._browse,
                      bg=P["border"], fg=P["text"], relief=tk.FLAT,
                      width=2, cursor="hand2").pack(side=tk.LEFT)
        row("CSV", _csv_row)

        self._csv_status_lbl = tk.Label(side, text="", bg=P["panel"],
                                        fg=P["subtext"], font=("Segoe UI", 8),
                                        wraplength=270, justify="left")
        self._csv_status_lbl.pack(anchor="w", padx=14)
        self._csv_var.trace_add("write", lambda *_: self._actualizar_csv_status())

        # Estado, sistema
        self._est_var = tk.StringVar(value="1000000000")
        self._sys_var = tk.StringVar(value="ABCDEFGHIJ")
        row("Estado", lambda f: tk.Entry(
            f, textvariable=self._est_var, width=20,
            bg=P["card"], fg=P["text"], insertbackground=P["text"],
            relief=tk.FLAT, font=("Consolas", 9)).pack(side=tk.LEFT))
        row("Sistema", lambda f: tk.Entry(
            f, textvariable=self._sys_var, width=20,
            bg=P["card"], fg=P["text"], insertbackground=P["text"],
            relief=tk.FLAT, font=("Consolas", 9)).pack(side=tk.LEFT))

        tk.Frame(side, height=1, bg=P["border"]).pack(fill=tk.X, padx=8, pady=6)

        # Alcance / Mecanismo
        self._alc_var = tk.StringVar(value="ABCDEFGHIJ")
        self._mec_var = tk.StringVar(value="ABCDEFGHIJ")
        row("Alcance", lambda f: tk.Entry(
            f, textvariable=self._alc_var, width=20,
            bg=P["card"], fg=P["text"], insertbackground=P["text"],
            relief=tk.FLAT, font=("Consolas", 9)).pack(side=tk.LEFT))
        row("Mecanismo", lambda f: tk.Entry(
            f, textvariable=self._mec_var, width=20,
            bg=P["card"], fg=P["text"], insertbackground=P["text"],
            relief=tk.FLAT, font=("Consolas", 9)).pack(side=tk.LEFT))

        tk.Frame(side, height=1, bg=P["border"]).pack(fill=tk.X, padx=8, pady=6)

        # Opciones k
        self._k_var = tk.StringVar(value="2")
        row("k (particiones)", lambda f: ttk.Combobox(
            f, textvariable=self._k_var,
            values=["2","3","4","5"], state="readonly", width=6
        ).pack(side=tk.LEFT))

        # Estrategias
        self._use_geo  = tk.BooleanVar(value=True)
        self._use_qnod = tk.BooleanVar(value=True)
        chk_f = tk.Frame(side, bg=P["panel"])
        chk_f.pack(fill=tk.X, padx=12, pady=4)
        tk.Label(chk_f, text="Estrategias:", bg=P["panel"], fg=P["subtext"],
                 font=("Segoe UI", 8)).pack(anchor="w")
        chk_row = tk.Frame(chk_f, bg=P["panel"])
        chk_row.pack(anchor="w")
        tk.Checkbutton(chk_row, text="KGeoMIP", variable=self._use_geo,
                       bg=P["panel"], fg=P["text"], activebackground=P["panel"],
                       selectcolor=P["card"], font=("Segoe UI", 9)).pack(side=tk.LEFT)
        tk.Checkbutton(chk_row, text="KQNodes", variable=self._use_qnod,
                       bg=P["panel"], fg=P["text"], activebackground=P["panel"],
                       selectcolor=P["card"], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=8)

        tk.Frame(side, height=1, bg=P["border"]).pack(fill=tk.X, padx=8, pady=6)

        # Botones de acción
        def btn(parent, text, cmd, color=None):
            c = color or P["border"]
            tk.Button(parent, text=text, command=cmd,
                      bg=c, fg=P["text"],
                      activebackground=P["border"], activeforeground=P["text"],
                      relief=tk.FLAT, font=("Segoe UI", 9, "bold"),
                      padx=8, pady=6, cursor="hand2", width=24
                      ).pack(fill=tk.X, padx=12, pady=2)

        btn(side, "▶  Ejecutar análisis", self._run, P["accent"].replace("f7","a5") if True else P["accent"])
        btn(side, "▶  Ejecutar suite",    self._ejecutar_suite)
        btn(side, "📤  Exportar a Excel", self._exportar_a_excel)
        btn(side, "💾  Guardar CSV",       self._save)
        btn(side, "🗑  Limpiar",           self._clear)

        tk.Frame(side, height=1, bg=P["border"]).pack(fill=tk.X, padx=8, pady=6)
        tk.Label(side, text="📊 Completar plataformas", bg=P["panel"], fg=P["subtext"],
                 font=("Segoe UI", 8), cursor="hand2").pack(anchor="w", padx=14, pady=2)
        self.bind_all("<F5>", lambda e: self._run())

        # Inicializar status del CSV
        self.after(200, self._actualizar_csv_status)

    def _build_main(self, parent):
        P = PALETTE
        main = tk.Frame(parent, bg=P["bg"])
        main.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(4, 8), pady=8)

        # Metric cards row
        cards_f = tk.Frame(main, bg=P["bg"])
        cards_f.pack(fill=tk.X, pady=(0, 6))
        self._card_phi  = MetricCard(cards_f, "φ mínimo (KGeoMIP)")
        self._card_time = MetricCard(cards_f, "Tiempo total")
        self._card_part = MetricCard(cards_f, "Bipartición óptima")
        self._card_n    = MetricCard(cards_f, "Subsistema n")
        for c in [self._card_phi, self._card_time, self._card_part, self._card_n]:
            c.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=4, ipady=4)

        # Tabbed area
        nb = ttk.Notebook(main)
        nb.pack(fill=tk.BOTH, expand=True)
        self._nb = nb

        self._build_tab_analisis(nb)
        self._build_tab_suites(nb)
        self._build_tab_validacion(nb)
        self._build_tab_historial(nb)

    def _build_tab_analisis(self, nb):
        P = PALETTE
        frm = tk.Frame(nb, bg=P["bg"])
        nb.add(frm, text=" 🔬 Análisis ")

        pane = tk.PanedWindow(frm, orient=tk.HORIZONTAL, bg=P["bg"],
                              sashwidth=5, sashrelief=tk.FLAT)
        pane.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Log panel
        log_f = tk.Frame(pane, bg=P["bg"])
        tk.Label(log_f, text="Registro de ejecución", bg=P["bg"],
                 fg=P["subtext"], font=("Segoe UI", 8)).pack(anchor="w", padx=4)
        self._log = ColoredLog(log_f)
        vsb = tk.Scrollbar(log_f, command=self._log.yview, bg=P["panel"])
        self._log.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._log.pack(fill=tk.BOTH, expand=True, padx=(4, 0))
        pane.add(log_f, minsize=300)

        # Results panel
        res_f = tk.Frame(pane, bg=P["bg"])
        tk.Label(res_f, text="Resultados por estrategia", bg=P["bg"],
                 fg=P["subtext"], font=("Segoe UI", 8)).pack(anchor="w", padx=4)
        tree_f = tk.Frame(res_f, bg=P["bg"])
        tree_f.pack(fill=tk.BOTH, expand=True)
        self._tree = ResultsTable(tree_f)
        vsb2 = tk.Scrollbar(tree_f, command=self._tree.yview, bg=P["panel"])
        self._tree.configure(yscrollcommand=vsb2.set)
        vsb2.pack(side=tk.RIGHT, fill=tk.Y)
        self._tree.pack(fill=tk.BOTH, expand=True, padx=(0, 0))

        # Sub-notebook analysis tabs
        sub_nb = ttk.Notebook(res_f)
        sub_nb.pack(fill=tk.BOTH, expand=True, pady=(4, 0))
        self._sub_nb = sub_nb

        def _txt_tab(label):
            f = tk.Frame(sub_nb, bg=P["bg"])
            sub_nb.add(f, text=label)
            t = ColoredLog(f, wrap=tk.WORD)
            vsb = tk.Scrollbar(f, command=t.yview, bg=P["panel"])
            t.configure(yscrollcommand=vsb.set)
            vsb.pack(side=tk.RIGHT, fill=tk.Y)
            t.pack(fill=tk.BOTH, expand=True)
            return t

        self._tab_correctitud = _txt_tab(" ✔ Correctitud ")
        self._tab_complejidad = _txt_tab(" ⚙ Complejidad ")
        self._tab_eficiencia  = _txt_tab(" ⚡ Eficiencia ")
        self._tab_comparacion = _txt_tab(" ⚖ Comparación ")

        pane.add(res_f, minsize=400)

    def _build_tab_suites(self, nb):
        P = PALETTE
        frm = tk.Frame(nb, bg=P["bg"])
        nb.add(frm, text=" 🧪 Suites ")

        top = tk.Frame(frm, bg=P["panel"])
        top.pack(fill=tk.X, padx=6, pady=6)

        tk.Label(top, text="Suite:", bg=P["panel"], fg=P["subtext"],
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=8)
        self._suite_var = tk.StringVar(value=list(SUITES.keys())[0])
        ttk.Combobox(top, textvariable=self._suite_var,
                     values=list(SUITES.keys()), state="readonly", width=10
                     ).pack(side=tk.LEFT, padx=4)

        tk.Button(top, text="▶ Ejecutar suite", command=self._ejecutar_suite,
                  bg=P["accent"].replace("f7","a5") if True else P["accent"],
                  fg=P["text"], relief=tk.FLAT, font=("Segoe UI", 9, "bold"),
                  padx=10, pady=4, cursor="hand2").pack(side=tk.LEFT, padx=8)

        self._suite_pb = ttk.Progressbar(top, mode="determinate", length=200)
        self._suite_pb.pack(side=tk.LEFT, padx=8)
        self._suite_lbl = tk.StringVar(value="")
        tk.Label(top, textvariable=self._suite_lbl, bg=P["panel"],
                 fg=P["subtext"], font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=6)

        # Suite results tree
        cols = ("Prueba", "Alcance", "Mecanismo", "n", "φ Geo", "φ QN", "t(s)")
        suite_tree_f = tk.Frame(frm, bg=P["bg"])
        suite_tree_f.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        self._suite_tree = ttk.Treeview(suite_tree_f, columns=cols,
                                        show="headings", selectmode="browse")
        widths = [50, 160, 160, 40, 90, 90, 70]
        for col, w in zip(cols, widths):
            self._suite_tree.heading(col, text=col)
            self._suite_tree.column(col, width=w, minwidth=30)
        self._suite_tree.tag_configure("phi_zero",    foreground=P["green"])
        self._suite_tree.tag_configure("phi_nonzero", foreground=P["yellow"])
        self._suite_tree.tag_configure("error_row",   foreground=P["red"])
        vsb = tk.Scrollbar(suite_tree_f, command=self._suite_tree.yview, bg=P["panel"])
        self._suite_tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self._suite_tree.pack(fill=tk.BOTH, expand=True)

    def _build_tab_validacion(self, nb):
        P = PALETTE
        frm = tk.Frame(nb, bg=P["bg"])
        nb.add(frm, text=" ✅ Validación §5.2.2 ")

        btn_bar = tk.Frame(frm, bg=P["panel"])
        btn_bar.pack(fill=tk.X, padx=6, pady=6)

        def vbtn(text, cmd):
            tk.Button(btn_bar, text=text, command=cmd,
                      bg=P["border"], fg=P["text"], relief=tk.FLAT,
                      font=("Segoe UI", 9), padx=8, pady=4,
                      cursor="hand2").pack(side=tk.LEFT, padx=4, pady=4)

        vbtn("Generar casos", self._run_val_generar)
        vbtn("Validar KGeoMIP", self._run_val_kgeomip)
        vbtn("Validar KQNodes", self._run_val_kqnodes)
        vbtn("Validar ambas", self._run_val_ambas)

        val_f = tk.Frame(frm, bg=P["bg"])
        val_f.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        self._tab_validacion_txt = ColoredLog(val_f, wrap=tk.NONE)
        vsb = tk.Scrollbar(val_f, command=self._tab_validacion_txt.yview, bg=P["panel"])
        hsb = tk.Scrollbar(val_f, orient=tk.HORIZONTAL,
                           command=self._tab_validacion_txt.xview, bg=P["panel"])
        self._tab_validacion_txt.configure(yscrollcommand=vsb.set,
                                           xscrollcommand=hsb.set)
        self._tab_validacion_txt.tag_configure("info", foreground=P["blue"])
        self._tab_validacion_txt.tag_configure("ok",   foreground=P["green"])
        self._tab_validacion_txt.tag_configure("warn", foreground=P["yellow"])
        self._tab_validacion_txt.tag_configure("err",  foreground=P["red"])
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._tab_validacion_txt.pack(fill=tk.BOTH, expand=True)

    def _build_tab_historial(self, nb):
        P = PALETTE
        frm = tk.Frame(nb, bg=P["bg"])
        nb.add(frm, text=" 📋 Historial ")

        top = tk.Frame(frm, bg=P["panel"])
        top.pack(fill=tk.X, padx=6, pady=6)
        tk.Button(top, text="📤 Exportar historial",
                  command=self._exportar_historial_excel,
                  bg=P["border"], fg=P["text"], relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=8, pady=4,
                  cursor="hand2").pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="📊 Completar plataformas",
                  command=self._completar_plataformas,
                  bg=P["border"], fg=P["text"], relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=8, pady=4,
                  cursor="hand2").pack(side=tk.LEFT, padx=4)
        tk.Button(top, text="🗑 Limpiar historial",
                  command=self._limpiar_historial,
                  bg=P["border"], fg=P["text"], relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=8, pady=4).pack(side=tk.LEFT, padx=4)

        cols = ("#", "Alcance", "Mecanismo", "n", "φ Geo", "Bipartición", "φ QN", "t Geo(s)", "t QN(s)")
        hist_f = tk.Frame(frm, bg=P["bg"])
        hist_f.pack(fill=tk.BOTH, expand=True, padx=6, pady=(0, 6))
        self._hist_tree = ttk.Treeview(hist_f, columns=cols,
                                       show="headings", selectmode="browse")
        widths = [35, 160, 160, 40, 90, 200, 90, 80, 80]
        for col, w in zip(cols, widths):
            self._hist_tree.heading(col, text=col)
            self._hist_tree.column(col, width=w, minwidth=30)
        self._hist_tree.tag_configure("phi_zero",    foreground=P["green"])
        self._hist_tree.tag_configure("phi_nonzero", foreground=P["yellow"])
        vsb = tk.Scrollbar(hist_f, command=self._hist_tree.yview, bg=P["panel"])
        hsb = tk.Scrollbar(hist_f, orient=tk.HORIZONTAL,
                           command=self._hist_tree.xview, bg=P["panel"])
        self._hist_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self._hist_tree.pack(fill=tk.BOTH, expand=True)

    def _build_statusbar(self):
        P = PALETTE
        tk.Frame(self, height=1, bg=P["border"]).pack(fill=tk.X, padx=8, pady=(4, 0))
        bar = tk.Frame(self, bg=P["panel"])
        bar.pack(fill=tk.X, padx=8, pady=(0, 6))
        self._status = tk.StringVar(value="Listo")
        tk.Label(bar, textvariable=self._status, anchor="w", width=50,
                 bg=P["panel"], fg=P["subtext"], font=("Segoe UI", 9)
                 ).pack(side=tk.LEFT, padx=10, pady=5)
        self._pb = ttk.Progressbar(bar, mode="indeterminate", length=180)
        self._pb.pack(side=tk.LEFT, padx=6, pady=5)

    # ── Helpers UI ────────────────────────────────────────────────────────────

    def _poll(self):
        try:
            while True:
                msg = self._q.get_nowait()
                kind = msg["kind"]
                if kind == "log":
                    self._log.append(msg["text"], msg.get("tag", ""))
                elif kind == "status":
                    self._status.set(msg["text"])
                elif kind == "row":
                    self._tree.add_row(msg["values"])
                    self._results.append(msg["values"])
                elif kind == "done":
                    self._on_done(msg)
        except queue.Empty:
            pass
        self.after(80, self._poll)

    def _post(self, **kw):
        self._q.put(kw)

    def _on_predefinido(self, *_):
        key = self._pre_var.get()
        if key in PREDEFINIDOS:
            p = PREDEFINIDOS[key]
            self._sys_var.set(p["sistema"])
            self._est_var.set(p["estado"])
            self._csv_var.set(p["csv"])
            self._alc_var.set(p["sistema"])
            self._mec_var.set(p["sistema"])

    def _browse(self):
        path = filedialog.askopenfilename(
            title="Seleccionar CSV", filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if path:
            self._csv_var.set(path)

    def _clear(self):
        self._log.clear()
        self._tree.clear()
        self._results.clear()
        for w in [self._tab_correctitud, self._tab_complejidad,
                  self._tab_eficiencia, self._tab_comparacion]:
            w.set_content("")
        for card in [self._card_phi, self._card_time, self._card_part, self._card_n]:
            card.set("—")
        self._status.set("Listo")

    def _limpiar_historial(self):
        for item in self._hist_tree.get_children():
            self._hist_tree.delete(item)
        self._historial_analisis.clear()

    def _actualizar_csv_status(self, *_):
        path = self._csv_var.get().strip()
        if not path:
            self._csv_status_lbl.configure(text="", fg=PALETTE["subtext"])
            return
        self._csv_status_lbl.configure(text="⏳ Verificando…", fg=PALETTE["subtext"])
        threading.Thread(target=self._analizar_csv, args=(path,), daemon=True).start()

    def _analizar_csv(self, ruta):
        if not os.path.exists(ruta):
            self.after(0, lambda: self._csv_status_lbl.configure(
                text="⚠ CSV no encontrado — se usará TPM sintética",
                fg=PALETTE["yellow"]))
            return
        try:
            n_vars = n_filas = 0
            with open(ruta, encoding="utf-8", errors="replace") as f:
                for i, line in enumerate(f):
                    line = line.strip()
                    if not line:
                        continue
                    if i == 0:
                        parts = line.split(",")
                        n_vars = len(parts)
                        try:
                            float(parts[0]); n_filas = 1
                        except ValueError:
                            pass
                    else:
                        n_filas += 1
            esperado = 2**n_vars if n_vars > 0 else 0
            ok = (n_filas == esperado) and (esperado > 0)
            txt = (f"✓ CSV válido — {n_filas:,} estados × {n_vars} variables" if ok
                   else f"⚠ CSV cargado — {n_filas:,} filas (esperado {esperado:,})")
            fg = PALETTE["green"] if ok else PALETTE["yellow"]
            self.after(0, lambda t=txt, c=fg: self._csv_status_lbl.configure(text=t, fg=c))
        except Exception as e:
            msg = f"✗ Error leyendo CSV: {e}"
            self.after(0, lambda m=msg: self._csv_status_lbl.configure(
                text=m, fg=PALETTE["red"]))

    # ── Ejecución manual ──────────────────────────────────────────────────────

    def _run(self):
        if not MODULES_OK:
            messagebox.showerror("Módulos no disponibles", MODULES_ERROR); return
        self._clear()
        self._pb.start(12)
        self._status.set("Ejecutando análisis…")
        t = threading.Thread(target=self._worker, daemon=True)
        t.start()

    def _worker(self):
        import numpy as np
        sistema  = self._sys_var.get().strip().upper()
        estado   = self._est_var.get().strip()
        csv_path = self._csv_var.get().strip()
        alcance  = self._alc_var.get().strip().upper()
        mec      = self._mec_var.get().strip().upper()
        usar_geo  = self._use_geo.get()
        usar_qnod = self._use_qnod.get()

        self._ultimo_n_sistema = len(sistema)
        self._ultimo_sistema   = sistema
        t0  = time.time()
        sub = None

        try:
            if csv_path and os.path.exists(csv_path):
                self._post(kind="log", text=f"📂 CSV: {csv_path}", tag="info")
                sys_full = System.desde_csv(csv_path, estado)
                if sys_full.tpm.shape[0] != 2**sys_full.n:
                    n_filas, n_vars = sys_full.tpm.shape
                    if sys_full.n > 20 and n_filas == 2**20:
                        self._post(kind="log",
                                   text=f"  ⚠ CSV truncado a 2^20={n_filas:,} filas "
                                        f"(N={sys_full.n} > 20, límite de memoria)",
                                   tag="warn")
                    elif n_vars > 0 and n_filas == 2**n_vars:
                        estado_csv = (estado + '0'*n_vars)[:n_vars]
                        sys_full = System.desde_csv(csv_path, estado_csv)
                        self._post(kind="log",
                                   text=f"  ⚠ Estado corregido: '{estado}'→'{estado_csv}'",
                                   tag="warn")
                    else:
                        raise ValueError(f"CSV no reconocido: {sys_full.tpm.shape}")
                self._post(kind="status", text="Construyendo subsistema…")
                self._post(kind="log", text=f"🔧 Subsistema: alcance={alcance} mec={mec}", tag="info")
                sub = sys_full.construir_subsistema(list(alcance), list(mec))
                if sub.n != sub.tpm.shape[1]:
                    cols_elim = list(range(sub.n, sub.tpm.shape[1]))
                    if cols_elim:
                        sub = sub.marginalizar_columnas(cols_elim)
                if sub.n > 0 and sub.tpm.shape[0] != 2**sub.n:
                    self._post(kind="log",
                               text=f"  ⚠ Forma inusual: {sub.tpm.shape[0]} filas ≠ 2^{sub.n}",
                               tag="warn")
                if sub.n == 0:
                    self._post(kind="log",
                               text="  ⚠ Alcance ∩ Mecanismo = ∅ → φ = 0", tag="warn")
            else:
                n_sub  = min(len(alcance), len(mec))
                n_rows = 2**min(n_sub, 20)
                seed   = sum(ord(c) for c in estado + alcance + mec) % (2**31)
                rng    = np.random.default_rng(seed)
                tpm    = rng.random((n_rows, n_sub))
                est_s  = (estado + "0"*n_sub)[:n_sub]
                sub    = System(tpm, est_s, list(alcance[:n_sub]))
                self._post(kind="log",
                           text=f"⚡ CSV no encontrado → subsistema sintético n={n_sub}",
                           tag="warn")

            self._post(kind="log",
                       text=f"  n={sub.n}  |  etiquetas={''.join(sub.etiquetas)}  |  tpm={sub.tpm.shape}",
                       tag="info")
            geo_res, qnod_res = self._run_strategies(sub, alcance, mec, usar_geo, usar_qnod)

        except Exception as exc:
            import traceback
            self._post(kind="log", text=f"\n❌ ERROR: {exc}", tag="err")
            self._post(kind="log", text=traceback.format_exc(), tag="err")
            self._post(kind="status", text=f"Error: {exc}")
            self._post(kind="done", elapsed=time.time()-t0, geo=None, qnod=None,
                       sistema=sistema, alcance=alcance, mec=mec, sub=sub)
            return

        self._post(kind="done", elapsed=time.time()-t0,
                   geo=geo_res, qnod=qnod_res,
                   sistema=sistema, alcance=alcance, mec=mec, sub=sub)

    def _run_strategies(self, sub, alcance, mec, usar_geo, usar_qnod):
        geo_res = qnod_res = None
        k = int(self._k_var.get())
        TIMEOUT = 180

        if usar_geo and MODULES_OK:
            self._post(kind="status", text="KGeoMIP ejecutando…")
            self._post(kind="log", text="\n─── KGeoMIP ───", tag="head")
            t0_g = time.time()
            try:
                if k == 2:
                    geo = KGeoMIP(sub)
                else:
                    geo = KGeoMIPKPartition(sub, k=k)
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                    fut = ex.submit(geo.ejecutar)
                    try:
                        geo_res = fut.result(timeout=TIMEOUT)
                    except concurrent.futures.TimeoutError:
                        self._post(kind="log",
                                   text=f"  ⏱ KGeoMIP timeout ({TIMEOUT}s)", tag="warn")
                        geo_res = None
                t_g = time.time() - t0_g
                if geo_res:
                    geo_res["tiempo"] = t_g
                    p1, p2 = geo_res["biparticion"]
                    etqs = sub.etiquetas
                    s1 = "".join(etqs[i] for i in p1 if i < len(etqs))
                    s2 = "".join(etqs[i] for i in p2 if i < len(etqs))
                    phi = geo_res.get("phi", float("inf"))
                    self._post(kind="log", text=f"  φ = {phi:.8f}", tag="ok")
                    self._post(kind="log", text=f"  Bipartición: {{{s1}}} | {{{s2}}}", tag="ok")
                    self._post(kind="log", text=f"  Tiempo: {t_g:.3f}s", tag="info")
                    self._post(kind="row", values=("KGeoMIP", k, f"{{{s1}}}|{{{s2}}}",
                                                   f"{phi:.8f}", f"{t_g:.3f}"))
                    self._ultimo_res_geo   = geo_res
                    self._ultimo_res_geo_full = geo_res
                    self._ultimo_t_geo     = t_g
            except Exception as e:
                self._post(kind="log", text=f"  ✗ KGeoMIP error: {e}", tag="err")

        if usar_qnod and MODULES_OK:
            self._post(kind="status", text="KQNodes ejecutando…")
            self._post(kind="log", text="\n─── KQNodes ───", tag="head")
            t0_q = time.time()
            try:
                qn = KQNodes(sub)
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                    fut = ex.submit(qn.aplicar_estrategia)
                    try:
                        qnod_res = fut.result(timeout=TIMEOUT)
                    except concurrent.futures.TimeoutError:
                        self._post(kind="log",
                                   text=f"  ⏱ KQNodes timeout ({TIMEOUT}s)", tag="warn")
                        qnod_res = None
                t_q = time.time() - t0_q
                if qnod_res:
                    qnod_res["tiempo"] = t_q
                    etqs = sub.etiquetas
                    por_k = qnod_res.get("resultados_por_k", {})
                    for k_val in sorted(por_k.keys()):
                        rk = por_k[k_val]
                        if rk and rk.get("biparticion") is not None:
                            phi_k = rk.get("phi", float("inf"))
                            if phi_k < float("inf"):
                                parts = rk["biparticion"]
                                part_str = " | ".join(
                                    "{"+"".join(etqs[v] for v in p if v < len(etqs))+"}"
                                    for p in parts)
                                self._post(kind="log",
                                           text=f"  k={k_val}: φ={phi_k:.8f}  {part_str}",
                                           tag="ok")
                                self._post(kind="row",
                                           values=(f"KQNodes k={k_val}", k_val, part_str,
                                                   f"{phi_k:.8f}", f"{t_q:.3f}"))
                    self._ultimo_res_qn = qnod_res
                    self._ultimo_t_qn   = t_q
            except Exception as e:
                self._post(kind="log", text=f"  ✗ KQNodes error: {e}", tag="err")

        return geo_res, qnod_res

    def _on_done(self, msg):
        self._pb.stop()
        elapsed = msg.get("elapsed", 0)
        geo   = msg.get("geo")
        qnod  = msg.get("qnod")
        sub   = msg.get("sub")
        alc   = msg.get("alcance", "")
        mec   = msg.get("mec", "")
        n_sis = len(msg.get("sistema", ""))

        self._status.set(f"Completado en {elapsed:.2f}s")
        self._post(kind="log", text=f"\n✅ Total: {elapsed:.2f}s", tag="ok")

        if geo:
            phi = geo.get("phi", float("inf"))
            self._card_phi.set(f"{phi:.6f}",
                               PALETTE["green"] if phi < 1e-9 else PALETTE["yellow"])
        if elapsed > 0:
            self._card_time.set(f"{elapsed:.3f}s")
        if geo and sub:
            etqs = sub.etiquetas
            p1, p2 = geo.get("biparticion", ([], []))
            s1 = "".join(etqs[i] for i in p1 if i < len(etqs))
            s2 = "".join(etqs[i] for i in p2 if i < len(etqs))
            self._card_part.set(f"{{{s1}}}|{{{s2}}}")
        if sub:
            self._card_n.set(str(sub.n))

        if geo or qnod:
            self._ultimo_sub = sub
            self._generar_analisis(geo, qnod, sub,
                                   geo.get("tiempo", 0) if geo else 0,
                                   qnod.get("tiempo", 0) if qnod else 0)
            self._historial_analisis.append({
                "alcance": alc, "mecanismo": mec, "n_sistema": n_sis,
                "res_geo": geo, "res_geo_full": geo,
                "res_qn": qnod, "sub": sub,
                "t_geo": geo.get("tiempo", 0) if geo else 0,
                "t_qn": qnod.get("tiempo", 0) if qnod else 0,
            })
            self._actualizar_historial_tree()

    def _generar_analisis(self, geo_res, qnod_res, sub, t_geo, t_qn):
        if not sub:
            return
        n    = sub.n
        etqs = sub.etiquetas

        def lbl(idxs):
            return "".join(etqs[i] for i in idxs if i < len(etqs)) if etqs else ""

        lines_c, lines_x, lines_e, lines_cmp = [], [], [], []

        # ── Correctitud ──────────────────────────────────────────────────────
        lines_c.append("═"*55)
        lines_c.append("  CORRECTITUD  —  §5.2.1")
        lines_c.append("═"*55)
        if geo_res:
            phi  = geo_res.get("phi", float("inf"))
            p1, p2 = geo_res.get("biparticion", ([], []))
            lines_c.append(f"  KGeoMIP (k=2)")
            lines_c.append(f"    φ (MIP)      = {phi:.8f}")
            lines_c.append(f"    Bipartición  = {{{lbl(p1)}}} | {{{lbl(p2)}}}")
            lines_c.append(f"    Resultado    : {'φ=0 → sistema causalmente independiente' if phi < 1e-9 else 'φ>0 → causalidad integrada detectada'}")
        if qnod_res:
            por_k = qnod_res.get("resultados_por_k", {})
            for k_val in sorted(por_k.keys()):
                rk = por_k[k_val]
                if rk and rk.get("phi", float("inf")) < float("inf"):
                    phi_k = rk["phi"]
                    parts = rk.get("biparticion", [])
                    pstr  = " | ".join(f"{{{lbl(p)}}}" for p in parts)
                    lines_c.append(f"  KQNodes k={k_val}")
                    lines_c.append(f"    φ (MIP)      = {phi_k:.8f}")
                    lines_c.append(f"    Partición    = {pstr}")

        # ── Complejidad ───────────────────────────────────────────────────────
        lines_x.append("═"*55)
        lines_x.append("  COMPLEJIDAD  —  §5.2.3")
        lines_x.append("═"*55)
        lines_x.append(f"  n (subsistema)       = {n}")
        lines_x.append(f"  Estados              = 2^{n} = {2**n if n <= 20 else '> 1M'}")
        k_geo = 2
        lines_x.append(f"  KGeoMIP k=2  BFS     = O(n · 2^n)")
        for k_val in [2, 3, 4, 5]:
            try:
                s = stirling2(n, k_val)
                lines_x.append(f"  S({n},{k_val}) biparticiones   = {s:,}")
            except Exception:
                pass

        # ── Eficiencia ────────────────────────────────────────────────────────
        lines_e.append("═"*55)
        lines_e.append("  EFICIENCIA  —  §5.2.4")
        lines_e.append("═"*55)
        if t_geo > 0:
            lines_e.append(f"  KGeoMIP tiempo       = {t_geo:.4f}s")
        if t_qn > 0:
            lines_e.append(f"  KQNodes tiempo       = {t_qn:.4f}s")
        if t_geo > 0 and t_qn > 0 and t_qn > 0:
            ratio = t_qn / t_geo if t_geo > 0 else 0
            lines_e.append(f"  Speedup GEO/QN       = {ratio:.2f}x")

        # ── Comparación ───────────────────────────────────────────────────────
        lines_cmp.append("═"*55)
        lines_cmp.append("  COMPARACIÓN  KGeoMIP vs KQNodes")
        lines_cmp.append("═"*55)
        phi_g = geo_res.get("phi", float("inf"))  if geo_res  else None
        phi_q = None
        if qnod_res:
            por_k = qnod_res.get("resultados_por_k", {})
            rk2 = por_k.get(2)
            if rk2:
                phi_q = rk2.get("phi", float("inf"))
        if phi_g is not None:
            lines_cmp.append(f"  φ KGeoMIP = {phi_g:.8f}")
        if phi_q is not None:
            lines_cmp.append(f"  φ QNodes  = {phi_q:.8f}")
        if phi_g is not None and phi_q is not None:
            diff = abs(phi_g - phi_q)
            lines_cmp.append(f"  Diferencia = {diff:.2e}")
            lines_cmp.append(f"  {'OK: resultados equivalentes' if diff < 1e-4 else 'DIFERENCIA significativa'}")

        self._tab_correctitud.set_content("\n".join(lines_c))
        self._tab_complejidad.set_content("\n".join(lines_x))
        self._tab_eficiencia.set_content("\n".join(lines_e))
        self._tab_comparacion.set_content("\n".join(lines_cmp))

    # ── Suite ─────────────────────────────────────────────────────────────────

    def _ejecutar_suite(self):
        if not MODULES_OK:
            messagebox.showerror("Módulos no disponibles", MODULES_ERROR); return
        if self._suite_running:
            messagebox.showinfo("Suite activa", "Una suite ya está en ejecución."); return

        # Determinar suite desde pestaña activa o combo
        key = self._suite_var.get()
        if key not in SUITES:
            key = self._pre_var.get()
        if key not in SUITES:
            messagebox.showwarning("Sin suite", f"No hay suite definida para {key}"); return

        suite = SUITES[key]
        self._suite_running   = True
        self._modo_suite_activo = True
        for item in self._suite_tree.get_children():
            self._suite_tree.delete(item)
        self._nb.select(1)  # Suites tab

        threading.Thread(
            target=self._suite_worker,
            args=(suite, key),
            daemon=True
        ).start()

    def _suite_worker(self, suite, key):
        import numpy as np
        pruebas = suite["pruebas"]
        sistema = suite["sistema"]
        estado  = suite["estado"]
        csv_p   = suite["csv"]
        total   = len(pruebas)

        self.after(0, lambda: self._suite_pb.configure(maximum=total, value=0))
        resultados_csv = []

        for i, (alcance, mec) in enumerate(pruebas, 1):
            self.after(0, lambda i=i, t=total: (
                self._suite_pb.configure(value=i),
                self._suite_lbl.set(f"{i}/{t}")
            ))
            self._post(kind="status", text=f"Suite {key}: {i}/{total} — alc={alcance} mec={mec}")
            self._post(kind="log",
                       text=f"\n[{i:02d}/{total}] {alcance} / {mec}", tag="head")

            t0   = time.time()
            sub  = None
            geo_res = qnod_res = None

            try:
                if csv_p and os.path.exists(csv_p):
                    sys_full = System.desde_csv(csv_p, estado)
                    sub = sys_full.construir_subsistema(list(alcance), list(mec))
                    if sub.n != sub.tpm.shape[1]:
                        cols_elim = list(range(sub.n, sub.tpm.shape[1]))
                        if cols_elim:
                            sub = sub.marginalizar_columnas(cols_elim)
                else:
                    n_sub  = min(len(alcance), len(mec))
                    n_rows = 2**min(n_sub, 20)
                    seed   = sum(ord(c) for c in estado + alcance + mec) % (2**31)
                    rng    = np.random.default_rng(seed)
                    tpm    = rng.random((n_rows, n_sub))
                    est_s  = (estado + "0"*n_sub)[:n_sub]
                    sub    = System(tpm, est_s, list(alcance[:n_sub]))

                self._ultimo_sub       = sub
                self._ultimo_n_sistema = len(sistema)
                geo_res, qnod_res = self._run_strategies(
                    sub, alcance, mec,
                    self._use_geo.get(), self._use_qnod.get())

                if geo_res:
                    self._ultimo_res_geo      = geo_res
                    self._ultimo_res_geo_full = geo_res
                    self._ultimo_t_geo        = geo_res.get("tiempo", 0)
                if qnod_res:
                    self._ultimo_res_qn = qnod_res
                    self._ultimo_t_qn   = qnod_res.get("tiempo", 0)

                elapsed = time.time() - t0
                phi_g = geo_res.get("phi", float("inf"))  if geo_res  else float("inf")
                phi_q = float("inf")
                if qnod_res:
                    por_k = qnod_res.get("resultados_por_k", {})
                    rk2 = por_k.get(2)
                    if rk2:
                        phi_q = rk2.get("phi", float("inf"))

                tag = "phi_zero" if min(phi_g, phi_q) < 1e-9 else "phi_nonzero"
                vals = (i, alcance, mec, sub.n if sub else "?",
                        f"{phi_g:.6f}" if phi_g < float("inf") else "∞",
                        f"{phi_q:.6f}" if phi_q < float("inf") else "∞",
                        f"{elapsed:.2f}")
                self.after(0, lambda v=vals, tg=tag:
                    self._suite_tree.insert("", tk.END, values=v, tags=(tg,)))

                resultados_csv.append({
                    "prueba": i, "alcance": alcance, "mecanismo": mec,
                    "n_sub": sub.n if sub else 0,
                    "phi_geo": phi_g, "phi_qn": phi_q, "tiempo_s": elapsed
                })

                if csv_p:
                    self._exportar_a_excel_path(csv_p, alcance, mec, len(sistema))

                self._historial_analisis.append({
                    "alcance": alcance, "mecanismo": mec, "n_sistema": len(sistema),
                    "res_geo": geo_res, "res_geo_full": geo_res,
                    "res_qn": qnod_res, "sub": sub,
                    "t_geo": self._ultimo_t_geo, "t_qn": self._ultimo_t_qn,
                })
                self.after(0, self._actualizar_historial_tree)

            except Exception as exc:
                import traceback
                self._post(kind="log", text=f"  ✗ Error: {exc}", tag="err")
                vals = (i, alcance, mec, "?", "ERR", "ERR", "0")
                self.after(0, lambda v=vals:
                    self._suite_tree.insert("", tk.END, values=v, tags=("error_row",)))

        # Guardar CSV de suite
        os.makedirs("results", exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        suite_csv = f"results/suite_{key.replace('=','').lower()}_{ts}.csv"
        try:
            with open(suite_csv, "w", newline="", encoding="utf-8") as f:
                w = csv.DictWriter(f, fieldnames=["prueba","alcance","mecanismo",
                                                   "n_sub","phi_geo","phi_qn","tiempo_s"])
                w.writeheader(); w.writerows(resultados_csv)
            self._post(kind="log", text=f"\n✅ Suite {key} completada. CSV: {suite_csv}", tag="ok")
        except Exception as e:
            self._post(kind="log", text=f"  ⚠ No se pudo guardar CSV: {e}", tag="warn")

        self._post(kind="status", text=f"Suite {key} completada — {total} pruebas")
        self._suite_running     = False
        self._modo_suite_activo = False

    # ── Historial ─────────────────────────────────────────────────────────────

    def _actualizar_historial_tree(self):
        for item in self._hist_tree.get_children():
            self._hist_tree.delete(item)
        for idx, h in enumerate(self._historial_analisis, 1):
            geo = h.get("res_geo")
            qn  = h.get("res_qn")
            sub = h.get("sub")
            etqs = sub.etiquetas if sub else []
            phi_g = geo.get("phi", float("inf")) if geo else float("inf")
            phi_q = float("inf")
            if qn:
                rk2 = qn.get("resultados_por_k", {}).get(2)
                if rk2:
                    phi_q = rk2.get("phi", float("inf"))
            bip = ""
            if geo and sub:
                p1, p2 = geo.get("biparticion", ([], []))
                s1 = "".join(etqs[i] for i in p1 if i < len(etqs))
                s2 = "".join(etqs[i] for i in p2 if i < len(etqs))
                bip = f"{{{s1}}}|{{{s2}}}"
            tag = "phi_zero" if phi_g < 1e-9 else "phi_nonzero"
            vals = (idx, h.get("alcance",""), h.get("mecanismo",""),
                    sub.n if sub else "?",
                    f"{phi_g:.6f}" if phi_g < float("inf") else "∞",
                    bip,
                    f"{phi_q:.6f}" if phi_q < float("inf") else "∞",
                    f"{h.get('t_geo',0):.3f}", f"{h.get('t_qn',0):.3f}")
            self._hist_tree.insert("", tk.END, values=vals, tags=(tag,))

    def _exportar_historial_excel(self):
        if not self._historial_analisis:
            messagebox.showinfo("Sin datos", "El historial está vacío."); return
        path = filedialog.asksaveasfilename(
            title="Guardar historial",
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv"), ("Todos", "*.*")])
        if not path:
            return
        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["#","Alcance","Mecanismo","n","phi_geo","biparticion",
                            "phi_qn","t_geo","t_qn"])
                for idx, h in enumerate(self._historial_analisis, 1):
                    geo = h.get("res_geo")
                    qn  = h.get("res_qn")
                    sub = h.get("sub")
                    etqs = sub.etiquetas if sub else []
                    phi_g = geo.get("phi", float("inf")) if geo else float("inf")
                    phi_q = float("inf")
                    if qn:
                        rk2 = qn.get("resultados_por_k", {}).get(2)
                        if rk2:
                            phi_q = rk2.get("phi", float("inf"))
                    bip = ""
                    if geo and sub:
                        p1, p2 = geo.get("biparticion", ([],[]))
                        s1 = "".join(etqs[i] for i in p1 if i < len(etqs))
                        s2 = "".join(etqs[i] for i in p2 if i < len(etqs))
                        bip = f"{s1}|{s2}"
                    w.writerow([idx, h.get("alcance",""), h.get("mecanismo",""),
                                sub.n if sub else "",
                                f"{phi_g:.8f}" if phi_g < float("inf") else "",
                                bip,
                                f"{phi_q:.8f}" if phi_q < float("inf") else "",
                                f"{h.get('t_geo',0):.4f}",
                                f"{h.get('t_qn',0):.4f}"])
            messagebox.showinfo("Exportado", f"Historial guardado en:\n{path}")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ── Excel export ──────────────────────────────────────────────────────────

    def _exportar_a_excel(self):
        n_total = (self._ultimo_n_sistema or len(self._sys_var.get().strip()))
        hoja_nombre = HOJAS_EXCEL.get(n_total)
        if not hoja_nombre:
            messagebox.showwarning(
                "Excel",
                f"N={n_total} no tiene hoja configurada.\n"
                f"Disponibles: N={list(HOJAS_EXCEL.keys())}"); return

        rutas = [
            'DatosPruebas2026_1.xlsx',
            os.path.join(os.path.dirname(os.path.abspath(__file__)), 'DatosPruebas2026_1.xlsx'),
            '../DatosPruebas2026_1.xlsx',
        ]
        excel_path = next((r for r in rutas if os.path.exists(r)), None)
        if not excel_path:
            excel_path = filedialog.askopenfilename(
                title="Selecciona DatosPruebas2026_1.xlsx",
                filetypes=[("Excel files","*.xlsx"),("Todos","*.*")])
        if not excel_path:
            return

        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            if hoja_nombre not in wb.sheetnames:
                messagebox.showwarning("Excel",
                    f"Hoja '{hoja_nombre}' no encontrada.\nHojas: {', '.join(wb.sheetnames)}")
                wb.close(); return

            ws      = wb[hoja_nombre]
            alcance = self._alc_var.get().strip().upper()
            mec     = self._mec_var.get().strip().upper()
            fila_destino = None
            for row in ws.iter_rows(min_row=6, max_col=3, values_only=False):
                if (str(row[1].value or "").strip().upper() == alcance and
                        str(row[2].value or "").strip().upper() == mec):
                    fila_destino = row[1].row; break

            if fila_destino is None:
                messagebox.showwarning("Excel",
                    f"No se encontró Alcance={alcance} / Mecanismo={mec} en '{hoja_nombre}'")
                wb.close(); return

            COL = {'qnodes_k2_part':4,'qnodes_k2_perd':5,'qnodes_k2_t':6,
                   'geo_k2_part':7,'geo_k2_perd':8,'geo_k2_t':9,
                   'qnodes_k3_part':10,'qnodes_k3_perd':11,'qnodes_k3_t':12,
                   'geo_k3_part':13,'geo_k3_perd':14,'geo_k3_t':15,
                   'qnodes_k4_part':16,'qnodes_k4_perd':17,'qnodes_k4_t':18,
                   'geo_k4_part':19,'geo_k4_perd':20,'geo_k4_t':21,
                   'qnodes_k5_part':22,'qnodes_k5_perd':23,'qnodes_k5_t':24,
                   'geo_k5_part':25,'geo_k5_perd':26,'geo_k5_t':27}

            etqs = self._ultimo_sub.etiquetas if self._ultimo_sub else []
            def lbl(idxs):
                return "".join(etqs[i] for i in idxs if i < len(etqs)) if etqs else ""
            def w(key, val):
                if val is not None:
                    ws.cell(row=fila_destino, column=COL[key], value=val)

            if self._ultimo_res_geo:
                rg = self._ultimo_res_geo
                p1, p2 = rg["biparticion"]
                w("geo_k2_part", (lbl(p1) or "") + "|" + (lbl(p2) or ""))
                w("geo_k2_perd", round(float(rg["phi"]), 8))
                w("geo_k2_t",    round(float(self._ultimo_t_geo), 4))

            if self._ultimo_res_qn:
                por_k = self._ultimo_res_qn.get("resultados_por_k", {})
                for k_val in [3, 4, 5]:
                    rk = por_k.get(k_val)
                    if rk and rk.get("biparticion") and rk.get("phi", float("inf")) < float("inf"):
                        parts = rk["biparticion"]
                        ps = " | ".join("{"+lbl(p)+"}" for p in parts)
                        w(f"qnodes_k{k_val}_part", ps)
                        w(f"qnodes_k{k_val}_perd", round(float(rk["phi"]), 8))
                        w(f"qnodes_k{k_val}_t",    round(float(self._ultimo_t_qn), 4))

            wb.save(excel_path); wb.close()
            messagebox.showinfo("Excel exportado",
                f"✓ Datos escritos.\n  Hoja: {hoja_nombre}\n  Fila: {fila_destino}\n  {excel_path}")
        except Exception as e:
            import traceback
            messagebox.showerror("Error exportando", f"{e}\n\n{traceback.format_exc()[:500]}")

    def _exportar_a_excel_path(self, excel_path: str, alcance: str,
                               mecanismo: str, n_sistema: int = None):
        try:
            import openpyxl
        except ImportError:
            self._post(kind="log", text="  ✗ openpyxl no instalado", tag="err"); return

        n_total     = n_sistema or self._ultimo_n_sistema or len(self._sys_var.get().strip())
        hoja_nombre = HOJAS_EXCEL.get(n_total)
        if not hoja_nombre:
            self._post(kind="log", text=f"  ✗ N={n_total} sin hoja Excel", tag="warn"); return
        if not os.path.exists(excel_path):
            self._post(kind="log", text=f"  ✗ Archivo no encontrado: {excel_path}", tag="err"); return

        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = None
            for h in wb.sheetnames:
                if h.strip() == hoja_nombre.strip():
                    ws = wb[h]; hoja_nombre = h; break
            if ws is None:
                norm = hoja_nombre.lower().replace("-","").replace(" ","")
                alt  = next((h for h in wb.sheetnames
                             if h.lower().replace("-","").replace(" ","") == norm), None)
                if alt:
                    ws = wb[alt]; hoja_nombre = alt
                    self._post(kind="log", text=f"  Usando hoja: '{hoja_nombre}'", tag="warn")
                else:
                    wb.close()
                    self._post(kind="log", text=f"  ✗ Hoja '{hoja_nombre}' no encontrada", tag="err")
                    return

            alc_up = alcance.strip().upper()
            mec_up = mecanismo.strip().upper()
            fila_destino = None
            for row in ws.iter_rows(min_row=6, max_col=3, values_only=False):
                if (str(row[1].value or "").strip().upper() == alc_up and
                        str(row[2].value or "").strip().upper() == mec_up):
                    fila_destino = row[0].row; break

            if fila_destino is None:
                self._post(kind="log",
                           text=f"  ⚠ Fila no encontrada: {alc_up}/{mec_up}", tag="warn")
                wb.close(); return

            etqs = self._ultimo_sub.etiquetas if self._ultimo_sub else []
            def lbl(idxs):
                return "".join(etqs[v] for v in idxs if v < len(etqs)) if etqs else ""
            def wr(col, val):
                if val is not None and val != "":
                    ws.cell(row=fila_destino, column=col, value=val)

            res_geo = self._ultimo_res_geo
            res_qn  = self._ultimo_res_qn
            por_k   = res_qn.get("resultados_por_k", {}) if res_qn else {}

            # QNodes k=2 (derivada de k=3)
            rk3 = por_k.get(3)
            if rk3 and rk3.get("biparticion") and rk3.get("phi", float("inf")) < float("inf"):
                parts3 = rk3["biparticion"]
                if len(parts3) >= 2:
                    s1 = lbl(parts3[0])
                    s2 = "".join(lbl(p) for p in parts3[1:])
                    wr(4, f"{s1}|{s2}")
                    wr(5, round(float(rk3["phi"]), 8))
                    t_qn = res_qn.get("tiempo", self._ultimo_t_qn) if res_qn else 0
                    wr(6, round(float(t_qn), 4))

            # Geo k=2
            if res_geo and res_geo.get("biparticion"):
                p1, p2 = res_geo["biparticion"]
                s1 = lbl(p1); s2 = lbl(p2)
                part_g = f"{s1}|{s2}" if (s1 or s2) else "∅"
                wr(7, part_g)
                wr(8, round(float(res_geo["phi"]), 8))
                wr(9, round(float(self._ultimo_t_geo), 4))

            # QNodes k=3,4,5
            COL_QN = {3:10, 4:16, 5:22}
            for k_val in [3, 4, 5]:
                rk = por_k.get(k_val)
                if rk and rk.get("biparticion") and rk.get("phi", float("inf")) < float("inf"):
                    parts    = rk["biparticion"]
                    part_str = " | ".join("{"+lbl(p)+"}" for p in parts)
                    col_qn   = COL_QN[k_val]
                    wr(col_qn,     part_str)
                    wr(col_qn + 1, round(float(rk["phi"]), 8))
                    t_k = rk.get("tiempo", self._ultimo_t_qn / 3 if self._ultimo_t_qn > 0 else 0)
                    wr(col_qn + 2, round(float(t_k), 4))

            # Geo k=3,4,5
            res_geo_full = getattr(self, "_ultimo_res_geo_full", None)
            por_k_geo    = res_geo_full.get("resultados_por_k", {}) if res_geo_full else {}
            GEO_COL = {3:13, 4:19, 5:25}
            for k_val in [3, 4, 5]:
                rk = por_k_geo.get(k_val)
                if rk and rk.get("biparticion") and rk.get("phi", float("inf")) < float("inf"):
                    parts    = rk["biparticion"]
                    part_str = " | ".join("{"+lbl(p)+"}" for p in parts)
                    cs = GEO_COL[k_val]
                    wr(cs,     part_str)
                    wr(cs + 1, round(float(rk["phi"]), 8))
                    wr(cs + 2, round(float(rk.get("tiempo", 0)), 4))

            try:
                wb.save(excel_path); wb.close()
                self._post(kind="log", text=f"  ✓ Excel: {hoja_nombre} fila {fila_destino}", tag="ok")
            except PermissionError:
                dir_d    = os.path.dirname(os.path.abspath(excel_path))
                base     = os.path.splitext(os.path.basename(excel_path))[0]
                copia    = os.path.join(dir_d, f"{base}_EXPORT.xlsx")
                try:
                    wb.save(copia); wb.close()
                    self._post(kind="log", text=f"  ⚠ Excel abierto — guardado como: {copia}", tag="warn")
                except Exception as e2:
                    self._post(kind="log", text=f"  ✗ No se pudo guardar: {e2}", tag="err")

        except Exception as e:
            import traceback
            self._post(kind="log",
                       text=f"  ✗ Error Excel: {e}\n{traceback.format_exc()[:300]}", tag="err")

    # ── Validación §5.2.2 ─────────────────────────────────────────────────────

    def _val_append(self, text: str, tag: str = "") -> None:
        def _do():
            self._tab_validacion_txt.configure(state=tk.NORMAL)
            self._tab_validacion_txt.insert(tk.END, text + "\n", (tag,) if tag else ())
            self._tab_validacion_txt.see(tk.END)
            self._tab_validacion_txt.configure(state=tk.DISABLED)
        self.after(0, _do)

    def _run_val_subprocess(self, cmd_args: list, desc: str) -> None:
        import subprocess
        self._val_append("="*55, "info")
        self._val_append(f"  {desc}", "info")
        self._val_append(f"  $ {' '.join(cmd_args)}", "info")
        self._val_append("="*55, "info")
        self._nb.select(2)  # Validación tab

        def _run():
            try:
                proc = subprocess.Popen(
                    cmd_args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                    text=True, encoding="utf-8", errors="replace",
                    cwd=os.path.dirname(os.path.abspath(__file__)))
                for line in proc.stdout:
                    self._val_append(line.rstrip())
                proc.wait()
                tag = "ok" if proc.returncode == 0 else "warn"
                self._val_append(f"\n  Terminó con código {proc.returncode}", tag)
            except Exception as exc:
                self._val_append(f"  ERROR: {exc}", "err")

        threading.Thread(target=_run, daemon=True).start()

    def _run_val_generar(self):
        self._run_val_subprocess(
            [sys.executable, "tests/leer_excel_referencia.py"],
            "Generando casos de referencia desde PruebasIniciales.xlsx…")

    def _run_val_kgeomip(self):
        self._run_val_subprocess(
            [sys.executable, "tests/ejecutar_validacion.py", "--estrategia", "geomip"],
            "Validando KGeoMIP contra casos de referencia…")

    def _run_val_kqnodes(self):
        self._run_val_subprocess(
            [sys.executable, "tests/ejecutar_validacion.py", "--estrategia", "qnodes"],
            "Validando KQNodes (k=2 exhaustivo) contra casos de referencia…")

    def _run_val_ambas(self):
        self._run_val_subprocess(
            [sys.executable, "tests/ejecutar_validacion.py"],
            "Validando KGeoMIP + KQNodes (resumen comparativo)…")

    # ── Exportar plataformas ───────────────────────────────────────────────────

    def _completar_plataformas(self):
        import platform, multiprocessing
        try:
            import psutil
        except ImportError:
            messagebox.showwarning("psutil", "pip install psutil para usar esta función"); return
        excel_path = filedialog.askopenfilename(
            title="Selecciona DatosPruebas2026_1.xlsx",
            filetypes=[("Excel files","*.xlsx")])
        if not excel_path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            ws = wb["plataformas"]
            ram_gb = round(psutil.virtual_memory().total / (1024**3), 0)
            so     = f"{platform.system()} {platform.release()}"
            proc   = platform.processor() or platform.machine()
            freq   = ""
            try:
                fi = psutil.cpu_freq()
                if fi:
                    freq = f"{fi.max/1000:.2f} GHz"
            except Exception:
                pass
            ncpus = multiprocessing.cpu_count()
            ws["B3"] = proc
            ws["C3"] = f"{int(ram_gb)} GB"
            ws["D3"] = so
            ws["B4"] = f"{freq}  {ncpus} núcleos".strip()
            wb.save(excel_path); wb.close()
            messagebox.showinfo("Plataformas",
                f"Información escrita:\n  CPU: {proc}\n  RAM: {int(ram_gb)} GB\n"
                f"  S.O.: {so}\n  Freq: {freq}  {ncpus} núcleos")
        except Exception as e:
            messagebox.showerror("Error plataformas", str(e))

    # ── Guardar CSV ───────────────────────────────────────────────────────────

    def _save(self):
        if not self._results:
            messagebox.showinfo("Sin datos", "No hay resultados para guardar."); return
        os.makedirs("results", exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = f"results/resultado_manual_{ts}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["Estrategia","k","Partición","Pérdida (φ)","Tiempo (s)"])
            w.writerows(self._results)
        messagebox.showinfo("Guardado", f"Resultados exportados a:\n{path}")


if __name__ == "__main__":
    app = App()
    app.mainloop()
